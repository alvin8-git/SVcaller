#!/usr/bin/env python3
"""Build per-sample SVcaller HTML report using Jinja2. Callers: Manta+Delly+GRIDSS+Scramble+MELT+SvABA+STRling."""
import argparse, csv, gzip, json, re, sys
from datetime import date
from pathlib import Path
from jinja2 import Environment, FileSystemLoader


TEMPLATE_DIR = Path(__file__).parent.parent / "assets"
STR_RANGES_PATH = TEMPLATE_DIR / "str_disease_ranges.tsv"
CNV_SYNDROMES_PATH = TEMPLATE_DIR / "cnv_syndromes.tsv"

# Canonical autosomes + sex chrs for per-chr QC table
_CANONICAL = [f"chr{c}" for c in list(range(1, 23)) + ["X", "Y"]]
_LOW_COV_THRESHOLD = 15.0
_MAX_ARTIFACT_SIZE = 50_000_000  # chromosome-spanning calls — always artifacts
_LARGE_SV_SUPP_MIN = 2           # require multi-caller support for SVs >1 Mb


_SD_PAIR_POS_WINDOW = 50_000   # bp — start position proximity for SD pair detection
_SD_PAIR_SIZE_TOL   = 0.20     # fractional size difference tolerance

_GNOMAD_AF_THRESHOLD = 0.01   # SVs with population AF > 1% are common variation, not P/LP

# ACMG Secondary Findings v3.2 (2023) — genes where incidental findings require reporting
ACMG_SF_V32_GENES = frozenset({
    # Hereditary breast/ovarian cancer
    "BRCA1", "BRCA2", "PALB2",
    # Lynch / MMR
    "MLH1", "MSH2", "MSH6", "PMS2", "EPCAM",
    # Polyposis
    "APC", "MUTYH", "SMAD4", "BMPR1A",
    # Cancer syndromes
    "TP53", "STK11", "PTEN", "VHL", "MEN1", "RET",
    "RB1", "WT1", "FH", "FLCN", "TSC1", "TSC2",
    "CDH1", "DICER1", "NF1", "NF2", "SMARCB1", "LZTR1",
    # Paraganglioma / pheochromocytoma
    "SDHB", "SDHC", "SDHD", "SDHAF2", "SDHA", "MAX", "TMEM127",
    # HCM
    "MYBPC3", "MYH7", "TNNT2", "TNNI3", "TPM1", "MYL3", "MYL2",
    "ACTC1", "PRKAG2", "GLA", "LAMP2",
    # DCM / cardiomyopathy
    "LMNA", "SCN5A", "RBM20", "TNNC1",
    # ARVC
    "PKP2", "DSG2", "DSC2", "TMEM43", "DSP",
    # Channelopathies (LQTS / CPVT / Brugada)
    "KCNQ1", "KCNH2", "RYR2", "CASQ2", "TRDN",
    # Aortopathy / connective tissue
    "FBN1", "FBN2", "TGFBR1", "TGFBR2", "SMAD3",
    "ACTA2", "MYH11", "MYLK", "COL3A1", "LOX",
    # Familial hypercholesterolaemia
    "LDLR", "APOB", "PCSK9",
    # Malignant hyperthermia
    "RYR1", "CACNA1S",
    # Metabolic
    "HFE", "ATP7B", "SERPINA1",
    # Neurological
    "LRRK2", "GBA",
    # X-linked conditions
    "OTC", "RPGR",
    # Susceptibility / moderate-risk cancer
    "ATM", "CHEK2", "BRIP1", "RAD51C", "RAD51D", "MSH3",
})


def _dedup_pathogenic(rows: list) -> list:
    """Deduplicate P/LP calls: collapse same-(gene,svtype) groups, then remove
    reciprocal DUP+DEL pairs at the same locus (segmental duplication artifacts).

    SD artifact signature: same gene, one DUP + one DEL, starts within 50 kb,
    sizes within 20%. Both callers being fooled by the same SD copy-number
    ambiguity produces concordant but opposite-sign calls — not a real event.
    """
    from collections import defaultdict

    # Step 1: collapse same-(gene,svtype) groups → best-supported representative
    groups: dict = defaultdict(list)
    for r in rows:
        groups[(r["gene"], r["svtype"])].append(r)
    deduped = []
    for group in groups.values():
        group.sort(key=lambda r: (
            -(int(r["supp"]) if str(r["supp"]).isdigit() else 0),
            -(int(r["acmg_class"]) if str(r["acmg_class"]).isdigit() else 0),
        ))
        best = group[0]
        if len(group) > 1:
            best = dict(best)
            best["collapsed"] = len(group)
        deduped.append(best)

    # Step 2: remove reciprocal DUP+DEL pairs at the same locus.
    # Match by CHROMOSOME + POSITION, not gene name, because paired calls at the
    # same locus often get different primary gene annotations from AnnotSV.
    dups_all = [r for r in deduped if r["svtype"] == "DUP"]
    dels_all = [r for r in deduped if r["svtype"] == "DEL"]

    sd_artifact_ids: set = set()
    for dup in dups_all:
        for del_ in dels_all:
            if dup.get("chrom") != del_.get("chrom"):
                continue
            try:
                pos_diff  = abs(int(dup["start"]) - int(del_["start"]))
                dup_size  = dup.get("size_bp", 0)
                del_size  = del_.get("size_bp", 0)
                size_diff = abs(dup_size - del_size) / max(dup_size, del_size, 1)
            except (ValueError, TypeError):
                continue
            if pos_diff <= _SD_PAIR_POS_WINDOW and size_diff <= _SD_PAIR_SIZE_TOL:
                sd_artifact_ids.add(id(dup))
                sd_artifact_ids.add(id(del_))

    deduped = [r for r in deduped if id(r) not in sd_artifact_ids]
    deduped.sort(key=lambda r: (-int(r["acmg_class"]), r["gene"], r["svtype"]))
    return deduped


def _parse_supp_vec(info_str: str) -> dict:
    """Extract SUPP, SUPP_VEC and caller breakdown from Jasmine INFO field."""
    info_d = dict(
        (f.split("=", 1) if "=" in f else (f, ""))
        for f in info_str.split(";")
    )
    supp     = info_d.get("SUPP", "?")
    supp_vec = info_d.get("SUPP_VEC", "")
    callers  = []
    names    = ["Manta", "Delly", "GRIDSS", "Scramble", "MELT", "SvABA"]
    for i, bit in enumerate(supp_vec):
        if bit == "1" and i < len(names):
            callers.append(names[i])
    return {"supp": supp, "supp_vec": supp_vec, "callers": "/".join(callers) or "?"}


def _load_str_ranges() -> dict:
    """Load curated STR disease/repeat-range metadata keyed by locus_id."""
    ranges = {}
    try:
        with open(STR_RANGES_PATH) as fh:
            for row in csv.DictReader(fh, delimiter="\t"):
                lid = row.get("locus_id", "").strip()
                if lid:
                    ranges[lid] = row
    except FileNotFoundError:
        pass
    return ranges


# ---------------------------------------------------------------------------
# SV parsing
# ---------------------------------------------------------------------------

def parse_sv_summary(sv_tsv_path: str) -> list:
    """Count SVs per type; flag ACMG class 4/5 (Path/LP)."""
    counts: dict = {}
    pathogenic: dict = {}
    try:
        with open(sv_tsv_path) as fh:
            reader = csv.DictReader(fh, delimiter="\t")
            for row in reader:
                if row.get("Annotation_mode", "") != "full":
                    continue
                st = row.get("SV_type", row.get("SVTYPE", "UNK")).upper()
                counts[st] = counts.get(st, 0) + 1
                if str(row.get("ACMG_class", "")).strip() in ("4", "5"):
                    pathogenic[st] = pathogenic.get(st, 0) + 1
    except (FileNotFoundError, KeyError):
        pass
    return [{"svtype": k, "total": v, "high": pathogenic.get(k, 0)}
            for k, v in sorted(counts.items())]


def _sv_row_from_annotated(row: dict):
    """Build unified SV row dict from a full-mode AnnotSV TSV row; returns None if excluded."""
    try:
        size_bp = abs(int(row.get("SV_end", 0) or 0) - int(row.get("SV_start", 0) or 0))
    except (ValueError, TypeError):
        size_bp = 0
    if size_bp > _MAX_ARTIFACT_SIZE:
        return None
    svt       = row.get("SV_type", "").upper()
    gene_full = row.get("Gene_name", row.get("Gene", ""))
    primary_gene = gene_full.split(";")[0].strip() if gene_full else "—"
    all_genes    = [g.strip() for g in gene_full.split(";") if g.strip()] if gene_full else []
    supp   = _parse_supp_vec(row.get("INFO", ""))
    supp_n = int(supp["supp"]) if str(supp["supp"]).isdigit() else 0
    af_col = {"DUP": "B_gain_AFmax", "DEL": "B_loss_AFmax",
              "INS": "B_ins_AFmax",  "INV": "B_inv_AFmax"}.get(svt, "")
    af_raw = (row.get(af_col, "") or "") if af_col else ""
    try:
        pop_af = float(af_raw) if af_raw and af_raw not in (".", "") else 0.0
    except ValueError:
        pop_af = 0.0
    def _hit(k):
        v = (row.get(k, "") or "").strip(); return bool(v and v != ".")
    try:
        acmg_score = float(row.get("AnnotSV_ranking_score", "") or 0)
    except (ValueError, TypeError):
        acmg_score = 0.0
    return {
        "svtype":        svt,
        "chrom":         row.get("SV_chrom", ""),
        "start":         row.get("SV_start", ""),
        "end":           row.get("SV_end",   ""),
        "size_bp":       size_bp,
        "size":          "",   # filled after _fmt_size defined
        "gene":          primary_gene,
        "gene_full":     gene_full,
        "all_genes":     all_genes,
        "omim":          row.get("OMIM_morbid", "") or "",
        "omim_candidate":row.get("OMIM_morbid_candidate", "") or "",
        "acmg_class":    str(row.get("ACMG_class", "")).strip(),
        "acmg_score":    acmg_score,
        "inheritance":   row.get("OMIM_inheritance", "") or "",
        "callers":       supp["callers"],
        "supp_n":        supp_n,
        "supp_vec":      supp.get("supp_vec", ""),
        "pop_af":        f"{pop_af:.4f}" if pop_af > 0 else "",
        "sd_boundary":   _hit("SegDup_left")  and _hit("SegDup_right"),
        "enc_blacklist": _hit("ENCODE_blacklist_left") and _hit("ENCODE_blacklist_right"),
        "pon_hit":       "SV_PON" in (row.get("INFO", "") or ""),
        "gnomad_common": pop_af > _GNOMAD_AF_THRESHOLD,
        "is_acmg_sf":    any(g in ACMG_SF_V32_GENES for g in all_genes),
        "acmg_genes":    [g for g in all_genes if g in ACMG_SF_V32_GENES],
        "tier":          3,
    }


def parse_sv_from_vcf(vcf_path: str) -> list:
    """Fallback SV rows from the merged (Jasmine) VCF when no AnnotSV annotation exists.

    Produces the same row dict shape as `_sv_row_from_annotated`, with annotation-derived
    fields left blank. Used so the SV sheet/HTML are never empty when SVs were called but
    `--annotsv_db` was not supplied (e.g. SMN-only runs) — the merged VCF still holds the calls.
    """
    rows = []
    opener = gzip.open if vcf_path.endswith(".gz") else open
    try:
        with opener(vcf_path, "rt") as fh:
            for line in fh:
                if line.startswith("#") or not line.strip():
                    continue
                f = line.rstrip("\n").split("\t")
                if len(f) < 8:
                    continue
                chrom, pos, _info = f[0], f[1], f[7]
                info_d = {}
                for kv in _info.split(";"):
                    if "=" in kv:
                        k, v = kv.split("=", 1); info_d[k] = v
                try:
                    start = int(pos)
                except ValueError:
                    continue
                try:
                    end = int(info_d.get("END", "") or start)
                except ValueError:
                    end = start
                svlen_raw = info_d.get("SVLEN", "")
                try:
                    size_bp = abs(int(svlen_raw)) if svlen_raw not in ("", ".") else abs(end - start)
                except (ValueError, TypeError):
                    size_bp = abs(end - start)
                if size_bp > _MAX_ARTIFACT_SIZE:
                    continue
                supp   = _parse_supp_vec(_info)
                supp_n = int(supp["supp"]) if str(supp["supp"]).isdigit() else 0
                rows.append({
                    "svtype": (info_d.get("SVTYPE", "") or "").upper(),
                    "chrom": chrom, "start": str(start), "end": str(end),
                    "size_bp": size_bp, "size": _fmt_size(size_bp),
                    "gene": "—", "gene_full": "", "all_genes": [],
                    "omim": "", "omim_candidate": "",
                    "acmg_class": "", "acmg_score": 0.0, "inheritance": "",
                    "callers": supp["callers"], "supp_n": supp_n,
                    "supp_vec": supp.get("supp_vec", ""),
                    "pop_af": "", "sd_boundary": False, "enc_blacklist": False,
                    "pon_hit": "SV_PON" in _info, "gnomad_common": False,
                    "is_acmg_sf": False, "acmg_genes": [], "tier": 3,
                })
    except FileNotFoundError:
        pass
    return rows


def classify_sv_tiers(sv_tsv_path: str) -> tuple:
    """Classify all full-mode AnnotSV rows into three clinical tiers.

    Tier 1 — Clinically actionable (ACMG SF v3.2 gene):
        Gene in ACMG_SF_V32_GENES AND supp ≥ 2 AND no PON AND not gnomAD common.
    Tier 2 — Candidate / likely TP:
        OMIM morbid gene AND supp ≥ 2 AND no PON; OR supp ≥ 3 regardless of OMIM.
    Tier 3 — Remaining, ranked by caller count then AnnotSV score.

    Returns (tier1, tier2, tier3, all_rows).
    """
    all_rows = []
    try:
        with open(sv_tsv_path) as fh:
            reader = csv.DictReader(fh, delimiter="\t")
            for row in reader:
                if row.get("Annotation_mode", "") != "full":
                    continue
                r = _sv_row_from_annotated(row)
                if r is None:
                    continue
                all_rows.append(r)
    except (FileNotFoundError, KeyError):
        pass

    tier1, tier2, tier3 = [], [], []
    for r in all_rows:
        if (r["is_acmg_sf"] and r["supp_n"] >= 2
                and not r["pon_hit"] and not r["gnomad_common"]):
            r["tier"] = 1
            tier1.append(r)
        elif ((r["omim"] and r["supp_n"] >= 2 and not r["pon_hit"])
              or r["supp_n"] >= 3):
            r["tier"] = 2
            tier2.append(r)
        else:
            r["tier"] = 3
            tier3.append(r)

    for r in all_rows:
        r["size"] = _fmt_size(r["size_bp"])

    tier1.sort(key=lambda r: (-(int(r["acmg_class"]) if r["acmg_class"].isdigit() else 0), r["gene"]))
    tier2.sort(key=lambda r: (-r["supp_n"], -r["acmg_score"]))
    tier3.sort(key=lambda r: (-r["supp_n"], -r["acmg_score"]))
    return tier1, tier2, tier3, all_rows


# placeholder so old call-sites don't break during transition
def parse_sv_pathogenic(sv_tsv_path: str) -> list:
    """Return Class 4/5 SVs ≤10 Mb with caller support for the findings highlight."""
    rows = []
    try:
        with open(sv_tsv_path) as fh:
            reader = csv.DictReader(fh, delimiter="\t")
            for row in reader:
                if row.get("Annotation_mode", "") != "full":
                    continue
                cls = str(row.get("ACMG_class", "")).strip()
                if cls not in ("4", "5"):
                    continue
                try:
                    size_bp = abs(int(row.get("SV_end", 0) or 0) - int(row.get("SV_start", 0) or 0))
                except (ValueError, TypeError):
                    size_bp = 0
                gene = row.get("Gene_name", row.get("Gene", ""))
                primary_gene = gene.split(";")[0] if gene else "—"
                supp = _parse_supp_vec(row.get("INFO", ""))
                supp_n = int(supp["supp"]) if str(supp["supp"]).isdigit() else 0
                if size_bp > _MAX_ARTIFACT_SIZE:
                    continue   # chromosome-spanning: always artifact
                if size_bp > 1_000_000 and supp_n < _LARGE_SV_SUPP_MIN:
                    continue   # large SV with single-caller support: likely artifact

                # P2 — gnomAD-SV population AF annotation (soft): flag common variants
                svt = row.get("SV_type", "").upper()
                if svt in ("DUP",):
                    af_raw = row.get("B_gain_AFmax", "") or ""
                elif svt == "DEL":
                    af_raw = row.get("B_loss_AFmax", "") or ""
                elif svt == "INS":
                    af_raw = row.get("B_ins_AFmax", "") or ""
                elif svt == "INV":
                    af_raw = row.get("B_inv_AFmax", "") or ""
                else:
                    af_raw = ""
                try:
                    pop_af = float(af_raw) if af_raw and af_raw not in (".", "") else 0.0
                except ValueError:
                    pop_af = 0.0
                gnomad_common = pop_af > _GNOMAD_AF_THRESHOLD

                # P1 — SD boundary annotation (soft flag): both breakpoints in segdup
                sd_left  = (row.get("SegDup_left",  "") or "").strip()
                sd_right = (row.get("SegDup_right", "") or "").strip()
                both_sd  = bool(sd_left and sd_right
                                and sd_left  not in (".", "")
                                and sd_right not in (".", ""))

                # P4 — ENCODE blacklist annotation (soft flag): both breakpoints in blacklist
                enc_left  = (row.get("ENCODE_blacklist_left",  "") or "").strip()
                enc_right = (row.get("ENCODE_blacklist_right", "") or "").strip()
                both_enc  = bool(enc_left and enc_right
                                 and enc_left  not in (".", "")
                                 and enc_right not in (".", ""))

                rows.append({
                    "svtype":         row.get("SV_type", ""),
                    "chrom":          row.get("SV_chrom", ""),
                    "start":          row.get("SV_start", ""),
                    "end":            row.get("SV_end", ""),
                    "gene":           primary_gene,
                    "omim":           row.get("OMIM_morbid", ""),
                    "acmg_class":     cls,
                    "size":           _fmt_size(size_bp),
                    "size_bp":        size_bp,
                    "callers":        supp["callers"],
                    "supp":           supp["supp"],
                    "collapsed":      0,
                    "pop_af":         f"{pop_af:.4f}" if pop_af > 0 else "",
                    "sd_boundary":    both_sd,
                    "enc_blacklist":  both_enc,
                    "pon_hit":        "SV_PON" in (row.get("INFO", "") or ""),
                    "gnomad_common":  gnomad_common,
                })
    except (FileNotFoundError, KeyError, ValueError):
        pass
    return _dedup_pathogenic(rows)


def parse_top_svs(sv_tsv_path: str, n_omim: int = 15, n_other: int = 5) -> list:
    """Top annotated SVs: OMIM morbid genes pinned first, then top n_other by AnnotSV score."""
    rows = []
    try:
        with open(sv_tsv_path) as fh:
            reader = csv.DictReader(fh, delimiter="\t")
            for row in reader:
                if row.get("Annotation_mode", "") != "full":
                    continue
                start = row.get("SV_start", "") or ""
                end   = row.get("SV_end",   "") or ""
                try:
                    size_bp = abs(int(end) - int(start))
                except (ValueError, TypeError):
                    size_bp = 0
                if size_bp > 10_000_000:
                    continue   # skip chromosome-spanning annotation artifacts
                acmg_score_raw = row.get("AnnotSV_ranking_score", row.get("Ranking", ""))
                try:
                    acmg_score = float(acmg_score_raw)
                except (ValueError, TypeError):
                    acmg_score = 0.0
                gene_full    = row.get("Gene_name", row.get("Gene", ""))
                primary_gene = gene_full.split(";")[0] if gene_full else "—"
                acmg_class   = str(row.get("ACMG_class", "")).strip()
                gnomad_af    = row.get("GD_AF", row.get("gnomAD_SV_AF", row.get("GNOMAD_AF", "")))
                inh_mode     = row.get("OMIM_inheritance", row.get("Gene_inheritance", ""))
                supp = _parse_supp_vec(row.get("INFO", ""))
                rows.append({
                    "chrom": row.get("SV_chrom", row.get("Chr", "")),
                    "start": start, "end": end,
                    "svtype": row.get("SV_type", ""),
                    "size":   _fmt_size(size_bp),
                    "gene":   primary_gene,
                    "gene_full": gene_full,
                    "acmg":  acmg_score_raw,
                    "acmg_class": acmg_class,
                    "omim":  row.get("OMIM_morbid", row.get("OMIM", "")),
                    "gnomad_af": gnomad_af,
                    "inheritance": inh_mode,
                    "callers": supp["callers"],
                    "supp":    supp["supp"],
                    "_score": acmg_score,
                })
    except (FileNotFoundError, KeyError):
        pass
    omim_rows  = sorted([r for r in rows if r.get("omim")],      key=lambda x: x["_score"], reverse=True)
    other_rows = sorted([r for r in rows if not r.get("omim")], key=lambda x: x["_score"], reverse=True)
    result = omim_rows[:n_omim] + other_rows[:n_other]
    for r in result:
        del r["_score"]
    return result


def _fmt_size(bp: int) -> str:
    if bp >= 1_000_000:
        return f"{bp/1_000_000:.1f} Mb"
    if bp >= 1_000:
        return f"{bp/1_000:.1f} kb"
    return f"{bp} bp"


_CNV_SYNDROME_MIN_OVERLAP = 0.5  # CNV must cover >=50% of a syndrome critical region


def load_cnv_syndromes(path=CNV_SYNDROMES_PATH) -> list:
    """Load curated recurrent pathogenic CNV syndrome regions (GRCh38)."""
    syndromes = []
    try:
        with open(path) as fh:
            for line in fh:
                if line.startswith("#") or not line.strip():
                    continue
                p = line.rstrip("\n").split("\t")
                if len(p) < 5:
                    continue
                try:
                    start, end = int(p[1]), int(p[2])
                except ValueError:
                    continue
                syndromes.append({
                    "chrom": p[0], "start": start, "end": end,
                    "type":  p[3].upper(), "syndrome": p[4],
                    "note":  p[5] if len(p) > 5 else "",
                })
    except FileNotFoundError:
        pass
    return syndromes


def match_cnv_syndrome(chrom, start, end, svtype, syndromes,
                       min_overlap=_CNV_SYNDROME_MIN_OVERLAP):
    """Return the syndrome a CNV hits, or None.

    Requires same chromosome, compatible type (syndrome DEL/DUP/BOTH vs CNV DEL/DUP),
    and the CNV covering >= min_overlap of the syndrome's critical region.
    """
    svtype = (svtype or "").upper()
    for s in syndromes:
        if s["chrom"] != chrom:
            continue
        if s["type"] != "BOTH" and s["type"] != svtype:
            continue
        region = s["end"] - s["start"]
        if region <= 0:
            continue
        ov = min(end, s["end"]) - max(start, s["start"])
        if ov > 0 and ov / region >= min_overlap:
            return s
    return None


def build_known_diseases(str_loci: list, smn_tsv_path: str = "",
                         cnv_bed_path: str = "", cnv_syndromes: list = None) -> list:
    """Tier 1a: named disease diagnoses from STR (EH EXPANDED), SMN (SMA), and CNVs that
    overlap a known recurrent pathogenic-CNV syndrome region."""
    findings = []

    # STR EXPANDED loci
    for locus in str_loci:
        if locus.get("status") != "EXPANDED":
            continue
        findings.append({
            "source":   "STR",
            "disease":  locus.get("disease") or locus.get("locus", ""),
            "gene":     locus.get("locus", ""),
            "detail":   f"{locus.get('repunit','')} repeat, CN: {locus.get('repcn','')}",
            "status":   "EXPANDED",
            "severity": "affected",
        })

    # SMN1/2 copy number — read TSV directly, skip comment lines
    if smn_tsv_path and smn_tsv_path not in ("NO_FILE", "null", ""):
        try:
            with open(smn_tsv_path) as fh:
                lines = [l for l in fh if l.strip() and not l.startswith("#")]
            if len(lines) >= 2:
                header = lines[0].rstrip("\n").split("\t")
                for data_line in lines[1:]:
                    row = dict(zip(header, data_line.rstrip("\n").split("\t")))
                    try:
                        smn1 = int(row.get("SMN1_CN", row.get("smn1", "-1")) or -1)
                        smn2 = int(row.get("SMN2_CN", row.get("smn2", "-1")) or -1)
                    except (ValueError, TypeError):
                        continue
                    if smn1 <= 1:
                        findings.append({
                            "source":   "SMN",
                            "disease":  "Spinal Muscular Atrophy (SMA)",
                            "gene":     "SMN1/SMN2",
                            "detail":   f"SMN1 CN: {smn1}, SMN2 CN: {smn2}",
                            "status":   "AFFECTED" if smn1 == 0 else "CARRIER",
                            "severity": "affected" if smn1 == 0 else "carrier",
                        })
        except (FileNotFoundError, KeyError):
            pass

    # CNVs are surfaced as a "known disease" ONLY when they overlap a curated recurrent
    # pathogenic-CNV syndrome region (DiGeorge, Williams, PWS/AS, 16p11.2, ...). Generic
    # large CNVs are NOT diagnoses — they live in the CNV Summary top-10 list and the xlsx.
    if cnv_syndromes is None:
        cnv_syndromes = load_cnv_syndromes()
    if cnv_syndromes and cnv_bed_path and cnv_bed_path not in ("NO_FILE", "null", ""):
        try:
            with open(cnv_bed_path) as fh:
                for line in fh:
                    if line.startswith("#") or not line.strip():
                        continue
                    parts = line.strip().split("\t")
                    if len(parts) < 7:
                        continue
                    chrom, start, end = parts[0], parts[1], parts[2]
                    svtype     = parts[4].upper() if len(parts) > 4 else ""
                    confidence = parts[6].upper() if len(parts) > 6 else ""
                    try:
                        start_i, end_i = int(start), int(end)
                    except (ValueError, TypeError):
                        continue
                    syn = match_cnv_syndrome(chrom, start_i, end_i, svtype, cnv_syndromes)
                    if not syn:
                        continue
                    size_bp = abs(end_i - start_i)
                    size_str = (f"{size_bp/1_000_000:.1f} Mb" if size_bp >= 1_000_000
                                else f"{size_bp/1_000:.0f} kb")
                    note = f" — {syn['note']}" if syn.get("note") else ""
                    findings.append({
                        "source":   "CNV",
                        "disease":  syn["syndrome"],
                        "gene":     f"{chrom}:{start}-{end}",
                        "detail":   f"{svtype} {size_str}, {confidence} confidence{note}",
                        "status":   confidence,
                        "severity": "affected" if confidence in ("BOTH", "HIGH") else "carrier",
                    })
        except (FileNotFoundError, IndexError):
            pass

    return findings


def export_xls(all_sv_rows: list, cnv_bed_path: str, str_loci: list,
               smn_tsv_path: str, out_path: str) -> None:
    """Write per-sample Excel workbook with SV / CNV / STR / SMN sheets."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        print("openpyxl not available — skipping XLS export", file=sys.stderr)
        return
    FILLS = {
        1: PatternFill("solid", fgColor="FFD7D7"),
        2: PatternFill("solid", fgColor="FFF3CD"),
        3: PatternFill("solid", fgColor="EEF2FF"),
    }
    BOLD = Font(bold=True)

    def _hdr(ws, cols):
        ws.append(cols)
        for cell in ws[1]:
            cell.font = BOLD
            cell.alignment = Alignment(wrap_text=True)

    wb = openpyxl.Workbook()

    # SV sheet
    ws_sv = wb.active
    ws_sv.title = "SV"
    _hdr(ws_sv, ["Tier", "Chr", "Start", "End", "Type", "Size (bp)", "Gene",
                 "OMIM Morbid", "ACMG Class", "AnnotSV Score", "Callers",
                 "# Callers", "Pop AF", "SD Boundary", "PON Hit",
                 "gnomAD Common", "ACMG SF Gene"])
    for r in sorted(all_sv_rows, key=lambda x: x["tier"]):
        ws_sv.append([
            f"Tier {r['tier']}", r["chrom"], r["start"], r["end"],
            r["svtype"], r["size_bp"], r["gene"], r["omim"] or "",
            r["acmg_class"], r["acmg_score"], r["callers"], r["supp_n"],
            r["pop_af"] or "", "Yes" if r["sd_boundary"] else "",
            "Yes" if r["pon_hit"] else "", "Yes" if r["gnomad_common"] else "",
            "Yes" if r["is_acmg_sf"] else "",
        ])
        fill = FILLS.get(r["tier"])
        if fill:
            for cell in ws_sv[ws_sv.max_row]:
                cell.fill = fill

    # CNV sheet
    ws_cnv = wb.create_sheet("CNV")
    _hdr(ws_cnv, ["Chr", "Start", "End", "Name", "Type", "Score", "Confidence"])
    try:
        with open(cnv_bed_path) as fh:
            for line in fh:
                if line.startswith("#") or not line.strip():
                    continue
                ws_cnv.append(line.strip().split("\t"))
    except FileNotFoundError:
        pass

    # STR sheet
    ws_str = wb.create_sheet("STR")
    _hdr(ws_str, ["Locus", "Disease", "Chr", "Pos", "Repeat Unit", "CN (EH)",
                  "Status", "Normal Max", "Pathogenic Min",
                  "Spanning Reads", "Flanking Reads", "INREPEAT"])
    for locus in str_loci:
        ws_str.append([
            locus.get("locus", ""), locus.get("disease", ""),
            locus.get("chrom", ""), locus.get("pos", ""),
            locus.get("repunit", ""), locus.get("repcn", ""),
            locus.get("status", ""), locus.get("normal_max", ""),
            locus.get("path_min", ""), locus.get("spanning", ""),
            locus.get("flanking", ""), "Yes" if locus.get("inrepeat") else "",
        ])

    # SMN sheet
    ws_smn = wb.create_sheet("SMN")
    _hdr(ws_smn, ["Sample", "SMN1_CN", "SMN2_CN", "SMN1_allele1",
                  "SMN1_allele2", "Interpretation"])
    if smn_tsv_path and smn_tsv_path not in ("NO_FILE", "null", ""):
        try:
            with open(smn_tsv_path) as fh:
                for row in csv.DictReader(fh, delimiter="\t"):
                    smn1 = row.get("SMN1_CN", row.get("smn1", ""))
                    smn2 = row.get("SMN2_CN", row.get("smn2", ""))
                    try:
                        interp = ("SMA Affected" if int(smn1) == 0
                                  else "SMA Carrier" if int(smn1) == 1
                                  else "Normal")
                    except (ValueError, TypeError):
                        interp = ""
                    ws_smn.append([
                        row.get("sample", row.get("Sample", "")),
                        smn1, smn2,
                        row.get("SMN1_allele1", ""), row.get("SMN1_allele2", ""),
                        interp,
                    ])
        except (FileNotFoundError, KeyError):
            pass

    wb.save(out_path)
    print(f"XLS report written to {out_path}", file=sys.stderr)


# ---------------------------------------------------------------------------
# CNV parsing
# ---------------------------------------------------------------------------

def parse_cnv_summary(cnv_bed_path: str) -> dict:
    """Count CNV calls by confidence tier from cnv_consensus.bed."""
    summary = {"total": 0, "del": 0, "dup": 0, "both": 0, "high": 0, "medium": 0}
    try:
        with open(cnv_bed_path) as fh:
            for line in fh:
                if line.startswith("#") or not line.strip():
                    continue
                parts = line.strip().split("\t")
                if len(parts) < 7:
                    continue
                svtype     = parts[4].upper() if len(parts) > 4 else ""
                confidence = parts[6].upper() if len(parts) > 6 else ""
                summary["total"] += 1
                if svtype == "DEL":
                    summary["del"] += 1
                elif svtype == "DUP":
                    summary["dup"] += 1
                if confidence == "BOTH":
                    summary["both"] += 1
                elif confidence == "HIGH":
                    summary["high"] += 1
                elif confidence == "MEDIUM":
                    summary["medium"] += 1
    except (FileNotFoundError, IndexError):
        pass
    return summary


def top_cnvs_by_size(cnv_bed_path: str, n: int = 10, cnv_syndromes: list = None) -> tuple:
    """Return (top_n_cnvs_by_size, total_count).

    CNVs have no truth set and no per-call gene annotation, so size is the most
    defensible ranking signal. Surface the n largest (most reviewable) with their
    confidence tier and any recurrent-syndrome-region hit; the full list is in the xlsx.
    """
    rows = []
    if cnv_syndromes is None:
        cnv_syndromes = load_cnv_syndromes()
    try:
        with open(cnv_bed_path) as fh:
            for line in fh:
                if line.startswith("#") or not line.strip():
                    continue
                p = line.strip().split("\t")
                if len(p) < 7:
                    continue
                try:
                    start_i, end_i = int(p[1]), int(p[2])
                except (ValueError, TypeError):
                    continue
                svtype  = p[4].upper() if len(p) > 4 else ""
                size_bp = abs(end_i - start_i)
                size_str = (f"{size_bp/1_000_000:.2f} Mb" if size_bp >= 1_000_000
                            else f"{size_bp/1_000:.0f} kb")
                syn = match_cnv_syndrome(p[0], start_i, end_i, svtype, cnv_syndromes)
                rows.append({
                    "pos":        f"{p[0]}:{p[1]}-{p[2]}",
                    "svtype":     svtype,
                    "size":       size_str,
                    "size_bp":    size_bp,
                    "cn":         p[3] if len(p) > 3 else "",
                    "caller":     p[5] if len(p) > 5 else "",
                    "confidence": p[6].upper() if len(p) > 6 else "",
                    "syndrome":   syn["syndrome"] if syn else "",
                })
    except (FileNotFoundError, IndexError):
        pass
    total = len(rows)
    rows.sort(key=lambda r: -r["size_bp"])
    return rows[:n], total


# ---------------------------------------------------------------------------
# STRling genome-wide STR parsing (P6)
# ---------------------------------------------------------------------------

def parse_strling(tsv_path: str, min_allele2: float = 100.0,
                  min_prob: float = 0.90, top_n: int = 50) -> list:
    """Parse STRling genotype TSV; return high-confidence expansion candidates.

    Clinical threshold: allele2_est >= 100 repeat units OR prob_expansion >= 0.90.
    Lower hits (~50 ru / 0.5 prob) are common polymorphisms — excluded here to
    reduce noise for clinical reporting.
    """
    rows = []
    if not tsv_path or tsv_path in ("NO_STRLING", "NO_FILE", "null"):
        return rows
    try:
        with open(tsv_path) as fh:
            reader = csv.DictReader(fh, delimiter="\t")
            for row in reader:
                try:
                    a2   = float(row.get("allele2_est", 0) or 0)
                    prob = float(row.get("prob_expansion", 0) or 0)
                except (ValueError, TypeError):
                    continue
                if a2 < min_allele2 and prob < min_prob:
                    continue
                ru  = row.get("repeatunit", row.get("repeat_unit", ""))
                rows.append({
                    "chrom":      row.get("#chrom", row.get("chrom", "")),
                    "left":       row.get("left",  ""),
                    "right":      row.get("right", ""),
                    "repeatunit": ru,
                    "allele1":    row.get("allele1_est", ""),
                    "allele2":    f"{a2:.0f}",
                    "spanning":   row.get("spanning", ""),
                    "flanking":   row.get("flanking", ""),
                    "prob":       f"{prob:.2f}",
                })
    except (FileNotFoundError, KeyError):
        pass
    rows.sort(key=lambda r: float(r["allele2"]) if r["allele2"] else 0, reverse=True)
    return rows[:top_n]


def build_str_consensus(str_loci: list, strling_loci: list) -> tuple:
    """Merge ExpansionHunter catalog loci + STRling corroboration into clinical STR table.

    Clinical report shows ONLY the 32 EH catalog loci (known disease associations).
    STRling is used exclusively for:
      (a) corroborating EH hits at the same coordinate (±500 bp) → tags source "Both"
      (b) counting novel high-confidence expansions for the summary line

    Novel STRling loci (no EH catalog match) are NOT shown in the main table —
    they lack gene/disease/threshold context required for clinical interpretation and
    belong in the companion TSV for research review.

    Returns (table, novel_count) where novel_count is the number of STRling-only hits.
    """
    _STATUS_RANK = {"EXPANDED": 0, "PREMUTATION": 1, "INTERMEDIATE": 2,
                    "UNKNOWN": 3, "NORMAL": 4}
    _COORD_WINDOW = 500

    result = []
    eh_positions: dict = {}   # chrom -> [(int_pos, result_index)]

    for locus in str_loci:
        idx = len(result)
        try:
            pos_int = int(locus["pos"])
        except (ValueError, TypeError):
            pos_int = 0
        entry = dict(locus)
        entry["source"]       = "EH"
        entry["allele2_est"]  = ""
        entry["strling_prob"] = ""
        result.append(entry)
        chrom = locus["chrom"]
        eh_positions.setdefault(chrom, []).append((pos_int, idx))

    novel_count = 0
    for sl in strling_loci:
        chrom = sl["chrom"]
        try:
            sl_pos = int(sl["left"])
        except (ValueError, TypeError):
            sl_pos = 0

        matched_idx = None
        for (eh_pos, idx) in eh_positions.get(chrom, []):
            if abs(eh_pos - sl_pos) <= _COORD_WINDOW:
                matched_idx = idx
                break

        if matched_idx is not None:
            # Corroborate existing EH entry
            result[matched_idx]["source"]       = "Both"
            result[matched_idx]["allele2_est"]  = sl["allele2"]
            result[matched_idx]["strling_prob"] = sl["prob"]
        else:
            # Novel locus — count only; not added to clinical table
            novel_count += 1

    def _sort_key(r):
        rank = _STATUS_RANK.get(r["status"], 3)
        try:
            nums  = [float(a) for a in r["repcn"].replace("/", " ").split() if a.replace(".", "").isdigit()]
            a_max = max(nums) if nums else 0.0
        except (ValueError, TypeError):
            a_max = 0.0
        return (rank, -a_max)

    result.sort(key=_sort_key)
    result = [r for r in result if r["status"] != "NORMAL"]
    return result, novel_count


# ---------------------------------------------------------------------------
# QC parsing
# ---------------------------------------------------------------------------

def parse_qc(coverage_path: str, metrics_path: str, flagstat_path: str = "",
             insert_size_path: str = "") -> dict:
    qc = {
        "mean_depth": "N/A", "dup_rate": "N/A",
        "mapped_pct": "N/A", "properly_paired_pct": "N/A",
        "total_reads": "N/A", "supplementary_pct": "N/A",
        "median_insert_size": "N/A", "mean_insert_size": "N/A",
        "insert_size_std": "N/A", "read_pairs": "N/A",
        "chr_coverage": [],    # list of {chrom, mean, flag}
        "low_cov_chrs": [],    # chrs below threshold
        "qc_pass": True,
        "qc_warnings": [],
    }
    _parse_mosdepth(coverage_path, qc)
    _parse_picard(metrics_path, qc)
    _parse_flagstat(flagstat_path, qc)
    _parse_insert_size(insert_size_path, qc)
    _compute_qc_status(qc)
    return qc


def _parse_mosdepth(summary_path: str, qc: dict) -> None:
    if not summary_path or summary_path in ("NO_FILE", "null"):
        return
    chr_cov = []
    try:
        with open(summary_path) as fh:
            for line in fh:
                if line.startswith("total\t"):
                    parts = line.strip().split("\t")
                    if len(parts) >= 4:
                        qc["mean_depth"] = f"{float(parts[3]):.1f}"
                    continue
                parts = line.strip().split("\t")
                if len(parts) < 4 or parts[0] == "chrom":
                    continue
                chrom = parts[0]
                if chrom not in _CANONICAL:
                    continue
                try:
                    mean = float(parts[3])
                except ValueError:
                    continue
                flag = mean < _LOW_COV_THRESHOLD and chrom not in ("chrY",)
                chr_cov.append({"chrom": chrom, "mean": f"{mean:.1f}", "flag": flag})
        qc["chr_coverage"] = chr_cov
        qc["low_cov_chrs"] = [c["chrom"] for c in chr_cov if c["flag"]]
    except (FileNotFoundError, ValueError):
        pass


def _parse_picard(metrics_path: str, qc: dict) -> None:
    if not metrics_path or metrics_path in ("NO_FILE", "null"):
        return
    try:
        with open(metrics_path) as fh:
            lines = fh.readlines()
        for i, line in enumerate(lines):
            stripped = line.strip()
            if not stripped or stripped.startswith("#"):
                continue
            if "PERCENT_DUPLICATION" in stripped:
                header = stripped.split("\t")
                data = lines[i + 1].strip().split("\t") if i + 1 < len(lines) else []
                d = dict(zip(header, data))
                pct = d.get("PERCENT_DUPLICATION", "")
                if pct:
                    qc["dup_rate"] = f"{float(pct) * 100:.2f}"
                break
    except (FileNotFoundError, ValueError, IndexError):
        pass


def _parse_insert_size(insert_size_path: str, qc: dict) -> None:
    if not insert_size_path or insert_size_path in ("NO_FILE", "null"):
        return
    try:
        with open(insert_size_path) as fh:
            lines = fh.readlines()
        for i, line in enumerate(lines):
            if "MEDIAN_INSERT_SIZE" in line:
                header = line.strip().split("\t")
                data = lines[i + 1].strip().split("\t") if i + 1 < len(lines) else []
                d = dict(zip(header, data))
                if d.get("MEDIAN_INSERT_SIZE"):
                    qc["median_insert_size"] = d["MEDIAN_INSERT_SIZE"]
                if d.get("MEAN_INSERT_SIZE"):
                    qc["mean_insert_size"] = f"{float(d['MEAN_INSERT_SIZE']):.0f}"
                if d.get("STANDARD_DEVIATION"):
                    qc["insert_size_std"] = f"{float(d['STANDARD_DEVIATION']):.0f}"
                if d.get("READ_PAIRS"):
                    n = int(d["READ_PAIRS"])
                    qc["read_pairs"] = f"{n/1_000_000:.1f}M"
                break
    except (FileNotFoundError, ValueError, IndexError):
        pass


def _parse_flagstat(flagstat_path: str, qc: dict) -> None:
    if not flagstat_path or flagstat_path in ("NO_FILE", "null"):
        return
    try:
        primary = 0
        primary_dup = 0
        supplementary = 0
        properly_paired = 0
        with open(flagstat_path) as fh:
            for line in fh:
                if " mapped (" in line and "primary" not in line:
                    m = re.search(r'\(([0-9.]+)%', line)
                    if m:
                        qc["mapped_pct"] = m.group(1)
                m = re.match(r'(\d+) \+ \d+ primary duplicates', line)
                if m:
                    primary_dup = int(m.group(1))
                m = re.match(r'(\d+) \+ \d+ primary\s*$', line)
                if m:
                    primary = int(m.group(1))
                m = re.match(r'(\d+) \+ \d+ supplementary', line)
                if m:
                    supplementary = int(m.group(1))
                m = re.match(r'(\d+) \+ \d+ properly paired', line)
                if m:
                    properly_paired = int(m.group(1))
        if primary > 0:
            qc["dup_rate"] = f"{primary_dup / primary * 100:.2f}"
            qc["total_reads"] = f"{primary/1_000_000:.1f}M"
            if properly_paired:
                qc["properly_paired_pct"] = f"{properly_paired / primary * 100:.1f}"
            total_with_supp = primary + supplementary
            if supplementary and total_with_supp > 0:
                qc["supplementary_pct"] = f"{supplementary / total_with_supp * 100:.2f}"
    except (FileNotFoundError, ValueError):
        pass


def _compute_qc_status(qc: dict) -> None:
    warnings = []
    try:
        depth = float(qc.get("mean_depth", 0))
        if depth < 20:
            warnings.append(f"Low coverage: {depth:.1f}x (threshold 20x)")
    except (ValueError, TypeError):
        pass
    try:
        dup = float(qc.get("dup_rate", 0))
        if dup > 20:
            warnings.append(f"High duplicate rate: {dup:.1f}% (threshold 20%)")
    except (ValueError, TypeError):
        pass
    if qc.get("low_cov_chrs"):
        warnings.append(f"Low coverage chromosomes: {', '.join(qc['low_cov_chrs'])}")
    try:
        mapped = float(qc.get("mapped_pct", 100))
        if mapped < 90:
            warnings.append(f"Low mapping rate: {mapped:.1f}% (threshold 90%)")
    except (ValueError, TypeError):
        pass
    qc["qc_warnings"] = warnings
    qc["qc_pass"] = len(warnings) == 0


# ---------------------------------------------------------------------------
# STR parsing
# ---------------------------------------------------------------------------

def parse_str_loci(str_vcf_path: str, str_ranges: dict = None) -> list:
    """Parse ExpansionHunter VCF; join with clinical range metadata."""
    if not str_vcf_path or str_vcf_path in ("NO_FILE", "NO_STR", "null"):
        return []
    if str_ranges is None:
        str_ranges = {}
    loci = []
    try:
        header_fields: list = []
        opener = gzip.open if str_vcf_path.endswith(".gz") else open
        with opener(str_vcf_path, "rt") as fh:
            for line in fh:
                if line.startswith("##"):
                    continue
                if line.startswith("#CHROM"):
                    header_fields = line.strip().lstrip("#").split("\t")
                    continue
                parts = line.strip().split("\t")
                if not parts or len(parts) < 8:
                    continue
                d = dict(zip(header_fields, parts))
                chrom = d.get("CHROM", "")
                pos   = d.get("POS", "")
                info  = d.get("INFO", "")
                fmt   = d.get("FORMAT", "")
                sample_val = parts[9] if len(parts) > 9 else ""
                info_d = dict(
                    (f.split("=", 1) if "=" in f else (f, ""))
                    for f in info.split(";")
                )
                # VARID holds the locus name in EH output (e.g. VARID=FMR1)
                locus_id = info_d.get("VARID") or info_d.get("REPID") or d.get("ID", ".")
                repunit  = info_d.get("RU") or info_d.get("REPUNIT") or "."
                ref_cn   = info_d.get("REF", ".")
                fmt_d = dict(zip(fmt.split(":"), sample_val.split(":")))
                repcn     = fmt_d.get("REPCN", ".")
                repci     = fmt_d.get("REPCI", ".")
                adsp      = fmt_d.get("ADSP", ".")  # spanning reads
                adfl      = fmt_d.get("ADFL", ".")  # flanking reads
                gt        = fmt_d.get("GT", ".")
                so        = fmt_d.get("SO", "")     # source: SPANNING/INREPEAT etc.
                # Join with clinical ranges
                meta = str_ranges.get(locus_id, {})
                status = _str_status(repcn, meta)
                loci.append({
                    "locus": locus_id,
                    "chrom": chrom, "pos": pos,
                    "repunit": repunit,
                    "ref_cn": ref_cn,
                    "repcn": repcn,
                    "repci": repci,
                    "spanning": adsp,
                    "flanking": adfl,
                    "gt": gt,
                    "inrepeat": "INREPEAT" in so,
                    "disease":  meta.get("disease", ""),
                    "inheritance": meta.get("inheritance", ""),
                    "normal_max":  meta.get("normal_max", ""),
                    "path_min":    meta.get("path_min", ""),
                    "notes":       meta.get("notes", ""),
                    "status": status,
                })
    except FileNotFoundError:
        pass
    return loci


def _str_status(repcn: str, meta: dict) -> str:
    """Classify STR call as NORMAL / PREMUTATION / EXPANDED / UNKNOWN."""
    if not meta:
        return "UNKNOWN"
    try:
        # repcn may be allele1/allele2 — take max allele
        alleles = [int(a) for a in repcn.split("/") if a.isdigit()]
        if not alleles:
            return "UNKNOWN"
        max_cn = max(alleles)
        path_min = meta.get("path_min", "")
        premut_max = meta.get("premut_max", "")
        normal_max = meta.get("normal_max", "")
        path_min_i = int(path_min) if str(path_min).isdigit() else None
        normal_max_i = int(normal_max) if str(normal_max).isdigit() else None
        premut_max_i = int(premut_max) if str(premut_max).isdigit() else None
        if path_min_i and max_cn >= path_min_i:
            return "EXPANDED"
        if premut_max_i and normal_max_i and normal_max_i < max_cn <= premut_max_i:
            return "PREMUTATION"
        if normal_max_i and max_cn <= normal_max_i:
            return "NORMAL"
        # No normal_max defined → classify as NORMAL when clearly below pathogenic threshold.
        # Only return INTERMEDIATE when the count is genuinely above a known normal ceiling.
        if normal_max_i is None and path_min_i and max_cn < path_min_i:
            return "NORMAL"
        return "INTERMEDIATE"
    except (ValueError, TypeError):
        return "UNKNOWN"


# ---------------------------------------------------------------------------
# Benchmark parsing
# ---------------------------------------------------------------------------

def parse_benchmark(json_path: str) -> list:
    """Parse Truvari summary.json — returns list with precision/recall/f1/tp/fp/fn."""
    try:
        with open(json_path) as fh:
            data = json.load(fh)
        if "precision" in data:
            return [{"svtype": "Overall",
                     "precision": data.get("precision") or 0,
                     "recall":    data.get("recall")    or 0,
                     "f1":        data.get("f1")        or 0,
                     "tp":        data.get("TP-comp")   or 0,
                     "fp":        data.get("FP")        or 0,
                     "fn":        data.get("FN")        or 0,
                     "base_cnt":  data.get("base cnt")  or 0,
                     "comp_cnt":  data.get("comp cnt")  or 0}]
        return [{"svtype": k,
                 "precision": v.get("precision") or 0,
                 "recall":    v.get("recall")    or 0,
                 "f1":        v.get("f1")        or 0,
                 "tp": 0, "fp": 0, "fn": 0, "base_cnt": 0, "comp_cnt": 0}
                for k, v in data.items() if isinstance(v, dict)]
    except (FileNotFoundError, json.JSONDecodeError):
        return []


def parse_benchmark_sizebin(json_path: str) -> list:
    """Parse per-size-bin Truvari sizebin JSON."""
    try:
        with open(json_path) as fh:
            data = json.load(fh)
        return [{"bin": k,
                 "precision": v.get("precision") or 0,
                 "recall":    v.get("recall")    or 0,
                 "f1":        v.get("f1")        or 0}
                for k, v in data.items()]
    except (FileNotFoundError, json.JSONDecodeError):
        return []


# ---------------------------------------------------------------------------
# Render
# ---------------------------------------------------------------------------

def _inline_bootstrap_css() -> str:
    """Return Bootstrap 5.3.0 CSS as a string for self-contained offline HTML.

    Reads from assets/bootstrap.min.css if pre-bundled (e.g. in Docker image),
    otherwise fetches from CDN at render time with a 15s timeout.
    Falls back to empty string — report still renders with custom CSS.
    """
    bundled = TEMPLATE_DIR / "bootstrap.min.css"
    if bundled.exists():
        return bundled.read_text()
    try:
        import urllib.request
        with urllib.request.urlopen(
            "https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css",
            timeout=15,
        ) as resp:
            css = resp.read().decode("utf-8")
        try:
            bundled.write_text(css)
        except OSError:
            pass
        return css
    except Exception as exc:
        print(f"WARNING: could not inline Bootstrap CSS ({exc}); report uses CDN fallback", file=sys.stderr)
        return ""


def _read_trait_tsv(path):
    """Read a one-record trait contract TSV (header line starts with '#')."""
    if not path:
        return None
    p = Path(path)
    if p.name in ("NO_FILE", "") or not p.exists():
        return None
    header, data = None, None
    for line in p.read_text().splitlines():
        if not line.strip():
            continue
        if line.startswith("#"):
            header = line.lstrip("#").split("\t")
        elif data is None:
            data = line.split("\t")
    if not header or not data:
        return None
    return dict(zip(header, data))


def build_cnv_traits_section(rh_status_path=None, amy1_path=None,
                             gst_null_path=None, lpa_kiv2_path=None):
    """Build the 'Blood Group & Copy-Number Traits' report card.

    Returns an HTML string (a Bootstrap card), or '' when no trait file is present.
    """
    rh   = _read_trait_tsv(rh_status_path)
    amy1 = _read_trait_tsv(amy1_path)
    gst  = _read_trait_tsv(gst_null_path)
    lpa  = _read_trait_tsv(lpa_kiv2_path)
    if not any([rh, amy1, gst, lpa]):
        return ""

    def esc(x):
        return (str(x).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))

    rows = []
    if rh:
        badge = "danger" if rh.get("Rh_status") == "neg" else "success"
        label = {"pos": "Rh(D) positive", "neg": "Rh(D) negative"}.get(
            rh.get("Rh_status"), rh.get("Rh_status", "unknown"))
        rows.append(
            '<tr><th>Rh(D) blood group</th>'
            '<td><span class="badge bg-{}">{}</span></td>'
            '<td>RHD copies: {} &middot; confidence: {}</td></tr>'.format(
                badge, esc(label), esc(rh.get("RHD_copies", "NA")),
                esc(rh.get("confidence", ""))))
    if amy1:
        rows.append(
            '<tr><th>AMY1 (salivary amylase)</th>'
            '<td>{} copies</td><td>{}</td></tr>'.format(
                esc(amy1.get("AMY1_copies", "NA")), esc(amy1.get("method", ""))))
    if gst:
        def gbadge(v):
            b = "warning" if v == "null" else "secondary"
            return '<span class="badge bg-{}">{}</span>'.format(b, esc(v))
        rows.append(
            '<tr><th>GSTM1 / GSTT1 (detox)</th>'
            '<td>GSTM1 {} &middot; GSTT1 {}</td>'
            '<td>null = homozygous gene deletion</td></tr>'.format(
                gbadge(gst.get("GSTM1", "unknown")), gbadge(gst.get("GSTT1", "unknown"))))
    if lpa:
        rows.append(
            '<tr><th>Lp(a) — LPA KIV-2</th>'
            '<td>{} repeat copies</td><td>{}</td></tr>'.format(
                esc(lpa.get("KIV2_copies", "NA")), esc(lpa.get("method", ""))))

    return (
        '<div class="card section-card">'
        '<div class="card-header"><h5>Blood Group &amp; Copy-Number Traits</h5></div>'
        '<div class="card-body">'
        '<table class="table table-sm table-bordered mb-1">'
        '<thead><tr><th>Trait</th><th>Result</th><th>Detail</th></tr></thead>'
        '<tbody>' + "".join(rows) + '</tbody></table>'
        '<p class="text-muted small mb-0">Targeted normalized read-depth estimates '
        '(RHD/AMY1/GSTM1/GSTT1/LPA-KIV2); CNV consensus used only to corroborate '
        'called deletions.</p>'
        '</div></div>'
    )


def render_report(sample_id: str, smn_html_path: str, cnv_bed_path: str,
                  sv_tsv_path: str, circos_svg_path: str, out_path: str,
                  pipeline_version: str = "1.0.0",
                  benchmark_json: str = None,
                  sizebin_json: str = None,
                  benchmark_v5q_json: str = None,
                  sizebin_v5q_json: str = None,
                  coverage_path: str = None,
                  metrics_path: str = None,
                  flagstat_path: str = None,
                  insert_size_path: str = None,
                  str_vcf_path: str = None,
                  strling_tsv_path: str = None,
                  smn_tsv_path: str = None,
                  sv_vcf_path: str = None,
                  rh_status_path: str = None,
                  amy1_path: str = None,
                  gst_null_path: str = None,
                  lpa_kiv2_path: str = None) -> None:
    env = Environment(loader=FileSystemLoader(str(TEMPLATE_DIR)))
    template = env.get_template("report_template.html")

    str_ranges = _load_str_ranges()

    bootstrap_css     = _inline_bootstrap_css()
    smn_html          = Path(smn_html_path).read_text()
    circos_svg_inline = Path(circos_svg_path).read_text()
    sv_summary        = parse_sv_summary(sv_tsv_path)
    sv_tier1, sv_tier2, sv_tier3, sv_all = classify_sv_tiers(sv_tsv_path)

    # Fallback: AnnotSV TSV is empty (e.g. run without --annotsv_db) but SVs were
    # called. Populate from the merged VCF so the SV sheet/HTML are never blank while
    # thousands of calls exist. Rows are unannotated (no gene/ACMG), all Tier 3.
    sv_unannotated = False
    if (not sv_all and sv_vcf_path
            and Path(sv_vcf_path).name not in ("NO_FILE", "NO_SV", "")
            and Path(sv_vcf_path).exists()):
        sv_all = parse_sv_from_vcf(sv_vcf_path)
        if sv_all:
            sv_unannotated = True
            from collections import Counter
            sv_all.sort(key=lambda r: (-r["supp_n"], -r["size_bp"]))
            sv_tier3 = sv_all
            counts = Counter(r["svtype"] for r in sv_all if r["svtype"])
            sv_summary = [{"svtype": k, "total": v, "high": 0}
                          for k, v in sorted(counts.items())]

    _TIER_HTML_MAX = 10
    sv_tier2_total = len(sv_tier2); sv_tier2 = sv_tier2[:_TIER_HTML_MAX]
    sv_tier3_total = len(sv_tier3); sv_tier3 = sv_tier3[:_TIER_HTML_MAX]
    cnv_summary       = parse_cnv_summary(cnv_bed_path)
    cnv_syndromes      = load_cnv_syndromes()
    cnv_top, cnv_total = top_cnvs_by_size(cnv_bed_path, n=10, cnv_syndromes=cnv_syndromes)
    cnv_traits_html    = build_cnv_traits_section(rh_status_path, amy1_path,
                                                  gst_null_path, lpa_kiv2_path)
    benchmark          = parse_benchmark(benchmark_json)           if benchmark_json      else None
    benchmark_bins     = parse_benchmark_sizebin(sizebin_json)     if sizebin_json        else None
    benchmark_v5q      = parse_benchmark(benchmark_v5q_json)       if benchmark_v5q_json  else None
    benchmark_bins_v5q = parse_benchmark_sizebin(sizebin_v5q_json) if sizebin_v5q_json    else None
    qc                = parse_qc(coverage_path or "", metrics_path or "",
                                 flagstat_path or "", insert_size_path or "")
    str_loci          = parse_str_loci(str_vcf_path or "", str_ranges)
    strling_loci      = parse_strling(strling_tsv_path or "")
    str_consensus, strling_novel_count = build_str_consensus(str_loci, strling_loci)
    known_diseases    = build_known_diseases(str_loci, smn_tsv_path or "", cnv_bed_path or "",
                                              cnv_syndromes=cnv_syndromes)
    cascade_flag      = any(
        not r.get("pon_hit") and not r.get("sd_boundary")
        for r in sv_tier1 + sv_tier2
    )

    # XLS export — written alongside HTML
    xls_path     = str(out_path).replace(".report.html", ".variants.xlsx")
    xls_filename = Path(xls_path).name
    export_xls(sv_all, cnv_bed_path, str_loci, smn_tsv_path or "", xls_path)

    html = template.render(
        bootstrap_css=bootstrap_css,
        sample_id=sample_id,
        pipeline_version=pipeline_version,
        run_date=date.today().isoformat(),
        qc=qc,
        sv_summary=sv_summary,
        sv_tier1=sv_tier1,
        sv_tier2=sv_tier2,
        sv_tier2_total=sv_tier2_total,
        sv_tier3=sv_tier3,
        sv_tier3_total=sv_tier3_total,
        known_diseases=known_diseases,
        xls_filename=xls_filename,
        cnv_summary=cnv_summary,
        cnv_top=cnv_top,
        cnv_total=cnv_total,
        cnv_traits_html=cnv_traits_html,
        smn_html=smn_html,
        circos_svg_inline=circos_svg_inline,
        benchmark=benchmark,
        benchmark_bins=benchmark_bins,
        benchmark_v5q=benchmark_v5q,
        benchmark_bins_v5q=benchmark_bins_v5q,
        str_consensus=str_consensus,
        strling_novel_count=strling_novel_count,
        cascade_flag=cascade_flag,
        sv_unannotated=sv_unannotated,
    )
    Path(out_path).write_text(html)
    print(f"HTML report written to {out_path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--sample",           required=True)
    parser.add_argument("--smn-html",         required=True)
    parser.add_argument("--cnv-bed",          required=True)
    parser.add_argument("--sv-tsv",           required=True)
    parser.add_argument("--circos-svg",       required=True)
    parser.add_argument("--out",              required=True)
    parser.add_argument("--pipeline-version", default="1.0.0")
    parser.add_argument("--benchmark",        default=None)
    parser.add_argument("--sizebin",          default=None)
    parser.add_argument("--benchmark-v5q",    default=None, dest="benchmark_v5q")
    parser.add_argument("--sizebin-v5q",      default=None, dest="sizebin_v5q")
    parser.add_argument("--coverage",         default=None)
    parser.add_argument("--metrics",          default=None)
    parser.add_argument("--flagstat",         default=None)
    parser.add_argument("--insert-size",      default=None, dest="insert_size")
    parser.add_argument("--str-vcf",          default=None)
    parser.add_argument("--strling-tsv",      default=None, dest="strling_tsv")
    parser.add_argument("--smn-tsv",          default=None, dest="smn_tsv")
    parser.add_argument("--sv-vcf",           default=None, dest="sv_vcf",
                        help="Merged SV VCF; fallback source for the SV sheet when the "
                             "AnnotSV TSV is empty (run without --annotsv_db).")
    parser.add_argument("--rh-status",        default=None, dest="rh_status")
    parser.add_argument("--amy1",             default=None, dest="amy1")
    parser.add_argument("--gst-null",         default=None, dest="gst_null")
    parser.add_argument("--lpa-kiv2",         default=None, dest="lpa_kiv2")
    args = parser.parse_args()
    render_report(
        sample_id=args.sample,
        smn_html_path=args.smn_html,
        cnv_bed_path=args.cnv_bed,
        sv_tsv_path=args.sv_tsv,
        circos_svg_path=args.circos_svg,
        out_path=args.out,
        pipeline_version=args.pipeline_version,
        benchmark_json=args.benchmark,
        sizebin_json=args.sizebin,
        benchmark_v5q_json=args.benchmark_v5q,
        sizebin_v5q_json=args.sizebin_v5q,
        coverage_path=args.coverage,
        metrics_path=args.metrics,
        flagstat_path=args.flagstat,
        insert_size_path=args.insert_size,
        str_vcf_path=args.str_vcf,
        strling_tsv_path=args.strling_tsv,
        smn_tsv_path=args.smn_tsv,
        sv_vcf_path=args.sv_vcf,
        rh_status_path=args.rh_status,
        amy1_path=args.amy1,
        gst_null_path=args.gst_null,
        lpa_kiv2_path=args.lpa_kiv2,
    )


if __name__ == "__main__":
    main()
